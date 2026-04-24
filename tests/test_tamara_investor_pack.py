"""Regression tests for the Tamara quarterly investor pack ingest pipeline.

Covers:
- Classifier: QUARTERLY_INVESTOR_PACK type recognized by filename + sheet names
- Classifier: ordering wins against INVESTOR_REPORT / COMPANY_PRESENTATION generics
- Parser: template validation, date range detection, pack-date resolution
- Parser: headline line-item extraction (KPIs + FS + Budget variance)
- Enrichment: MoM delta computation, graceful no-op when pack folder missing
"""

from __future__ import annotations

import importlib.util
import json
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from core.dataroom.classifier import DocumentType, classify_document


# ── Parser module is a script in ``scripts/`` — import it by file path ────────
# (Can't ``import scripts.ingest_tamara_investor_pack`` because ``scripts/`` has
# no ``__init__.py`` and we don't want to add one just for tests.)

_SCRIPT_PATH = Path(__file__).resolve().parent.parent / "scripts" / "ingest_tamara_investor_pack.py"
_spec = importlib.util.spec_from_file_location("ingest_tamara_investor_pack", _SCRIPT_PATH)
_parser_mod = importlib.util.module_from_spec(_spec)
sys.modules["ingest_tamara_investor_pack"] = _parser_mod
_spec.loader.exec_module(_parser_mod)


# ── Helpers ───────────────────────────────────────────────────────────────────


def _make_pack_workbook(path: Path, num_months: int = 3, last_month: datetime = None) -> None:
    """Create a minimal but valid 10-sheet investor pack xlsx at ``path``.

    ``last_month`` defaults to 2026-03-31. ``num_months`` controls how many
    consecutive months are written into the date-header row of each content
    sheet.
    """
    if last_month is None:
        last_month = datetime(2026, 3, 31)

    wb = openpyxl.Workbook()
    # Remove the default sheet that Workbook() creates
    wb.remove(wb.active)

    # Build the 10 required sheet names in canonical order
    expected = [
        "Investor Reporting Pack >>",
        "1. KPIs >>",
        "1.1 KPIs cons",
        "1.2 KPIs KSA",
        "1.3 KPIs UAE",
        "2. Financials>>",
        "2.1 FS Cons",
        "2.2 FS KSA",
        "2.3 FS UAE",
        "3. Performance v Budget",
    ]

    # Build a list of month-end dates going back num_months from last_month
    months = []
    yr, mo = last_month.year, last_month.month
    for i in range(num_months):
        months.insert(0, datetime(yr, mo, 28))  # safe day for any month
        mo -= 1
        if mo == 0:
            mo = 12
            yr -= 1

    for name in expected:
        ws = wb.create_sheet(name)
        if name.endswith(">>"):
            # Divider sheet — just a marker
            ws.cell(2, 2).value = name
            continue

        # Content sheet: write date headers + a few sample line items
        is_fs = name.startswith("2.")
        is_budget = name.startswith("3.")
        header_row = 5 if is_fs else 6

        # Title rows
        ws.cell(1, 2).value = "KPI Monthly Dashboard" if not is_fs else "Financial Statements"
        ws.cell(2, 2).value = name.split()[-1] + " - Monthly"

        # Section labels for Budget sheet
        if is_budget:
            ws.cell(4, 4).value = "Actuals"
            ws.cell(4, 21).value = "Budget - Base Case"
            ws.cell(4, 26).value = "Variance vs. Budget"

        # Write date headers
        for col_idx, dt in enumerate(months):
            ws.cell(header_row, 4 + col_idx).value = dt

        # Budget-specific: add variance columns AFTER the date cols
        if is_budget:
            # Budget block starts at col 21 with 3 months
            for col_idx in range(num_months):
                ws.cell(header_row, 21 + col_idx).value = months[col_idx]
            ws.cell(header_row, 21 + num_months).value = "YTD"
            ws.cell(header_row, 26).value = "Monthly Budget Variance (US$)"
            ws.cell(header_row, 27).value = "Monthly Budget Variance (%)"
            ws.cell(header_row, 28).value = "YTD Budget Variance (US$)"
            ws.cell(header_row, 29).value = "YTD Budget Variance (%)"
            # Also YTD on actuals side
            ws.cell(header_row, 4 + num_months).value = "YTD"

        # Write 3 line items starting at data_start_row
        data_start = header_row + 2  # skip header row + blank
        if name == "1.1 KPIs cons":
            ws.cell(data_start, 2).value = "# Annual active customers"
            for col_idx in range(num_months):
                ws.cell(data_start, 4 + col_idx).value = 5_000_000 + col_idx * 200_000
            ws.cell(data_start + 1, 2).value = "AOV"
            for col_idx in range(num_months):
                ws.cell(data_start + 1, 4 + col_idx).value = 150.0 - col_idx * 2
        elif name == "2.1 FS Cons":
            ws.cell(data_start, 2).value = "Total GMV"
            for col_idx in range(num_months):
                ws.cell(data_start, 4 + col_idx).value = 800_000_000 + col_idx * 50_000_000
            ws.cell(data_start + 1, 2).value = "Total Operating Revenue"
            for col_idx in range(num_months):
                ws.cell(data_start + 1, 4 + col_idx).value = 60_000_000 + col_idx * 5_000_000
            ws.cell(data_start + 2, 2).value = "Statutory Net Profit / (Loss)"
            for col_idx in range(num_months):
                ws.cell(data_start + 2, 4 + col_idx).value = 5_000_000 + col_idx * 500_000
        elif name == "3. Performance v Budget":
            ws.cell(data_start, 2).value = "Total GMV"
            # Actuals
            for col_idx in range(num_months):
                ws.cell(data_start, 4 + col_idx).value = 900_000_000
            ws.cell(data_start, 4 + num_months).value = 900_000_000 * num_months  # YTD
            # Budget
            for col_idx in range(num_months):
                ws.cell(data_start, 21 + col_idx).value = 850_000_000
            ws.cell(data_start, 21 + num_months).value = 850_000_000 * num_months  # YTD budget
            # Variance
            ws.cell(data_start, 27).value = 0.0588  # monthly variance %
            ws.cell(data_start, 29).value = 0.0588  # YTD variance %

    wb.save(path)


# ── Classifier tests ──────────────────────────────────────────────────────────


class TestClassifierQuarterlyPack:
    """QUARTERLY_INVESTOR_PACK must be detected reliably by filename + sheet rules."""

    def test_filename_quarterly_cons_pack(self):
        """The Q1-2026 Tamara file name format should classify as the new type."""
        result = classify_document(
            "/data/Tamara/dataroom/Financials/54.2.2 Management Financials/"
            "2. 1Q2026 Tamara Cons. Investor Pack.xlsx"
        )
        assert result == DocumentType.QUARTERLY_INVESTOR_PACK

    def test_filename_investor_monthly_reporting(self):
        """Prior monthly-cadence filenames should also classify as the new type."""
        for name in [
            "Investor Monthly Reporting_Nov'25.xlsx",
            "Investor Monthly Reporting_Dec'25.xlsx",
            "Investor Monthly Reporting_Jan'26.xlsx",
            "54.2.2.1 Investor Monthly Reporting_Nov'25.xlsx",
        ]:
            assert classify_document(name) == DocumentType.QUARTERLY_INVESTOR_PACK, (
                f"{name} should be QUARTERLY_INVESTOR_PACK"
            )

    def test_ordering_wins_over_investor_report(self):
        """The new rule sits BEFORE INVESTOR_REPORT — a file that could match
        both must land in QUARTERLY_INVESTOR_PACK."""
        # "Quarterly Report" alone → INVESTOR_REPORT.
        # "1Q2026 Investor Pack" → QUARTERLY_INVESTOR_PACK (new rule wins).
        assert (
            classify_document("hsbc_quarterly_report.pdf")
            == DocumentType.INVESTOR_REPORT
        )
        assert (
            classify_document("3Q2025 Cons Investor Pack.xlsx")
            == DocumentType.QUARTERLY_INVESTOR_PACK
        )

    def test_sheet_rule_fallback(self):
        """Even if filename doesn't match, the 10-sheet template signature
        (detected via sheet names) should classify correctly."""
        # Anonymous filename, but sheets match the template
        sheets = ["Cover", "1.1 KPIs cons", "2.1 FS Cons", "3. Performance v Budget"]
        result = classify_document(
            "/random/anon.xlsx", text_preview=None, sheet_names=sheets
        )
        assert result == DocumentType.QUARTERLY_INVESTOR_PACK

    def test_generic_investor_pack_not_misclassified(self):
        """A generic 'investor_pack' deck with no 'cons' or quarter prefix
        should still land in COMPANY_PRESENTATION (existing behaviour), not
        the new type. Rule is specific — requires 'Cons' OR quarterly prefix
        OR 'Monthly Reporting'."""
        # Only "investor pack" — matches COMPANY_PRESENTATION rule, not the
        # new QUARTERLY_INVESTOR_PACK rule (which requires stronger signal).
        assert (
            classify_document("Generic Investor Pack.pdf")
            == DocumentType.COMPANY_PRESENTATION
        )


# ── Parser tests ──────────────────────────────────────────────────────────────


class TestInvestorPackParser:
    """parse_investor_pack correctness + pack-date resolution."""

    def test_parses_minimal_template(self):
        """A minimal 10-sheet xlsx round-trips through the parser."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "test_pack.xlsx"
            _make_pack_workbook(xlsx, num_months=3)

            pack = _parser_mod.parse_investor_pack(xlsx)

            assert pack["meta"]["template_version"] == "tamara_investor_pack_v1"
            # All 3 country cuts present
            assert set(pack["kpis"].keys()) == {"cons", "ksa", "uae"}
            assert set(pack["financials"].keys()) == {"cons", "ksa", "uae"}
            # At least 1 line item extracted per cons sheet
            assert len(pack["kpis"]["cons"]["line_items"]) >= 1
            assert len(pack["financials"]["cons"]["line_items"]) >= 1

    def test_template_mismatch_raises(self):
        """Missing expected sheet → ValueError (not silent drift)."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "bad.xlsx"
            wb = openpyxl.Workbook()
            # Only has the default sheet, not the 10 expected ones
            wb.save(xlsx)

            with pytest.raises(ValueError, match="Template mismatch"):
                _parser_mod.parse_investor_pack(xlsx)

    def test_pack_date_cli_override(self):
        """--pack-date CLI arg wins over heuristic."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "test_pack.xlsx"
            _make_pack_workbook(xlsx, num_months=3)

            pack = _parser_mod.parse_investor_pack(xlsx, cli_pack_date="2026-05-01")

            assert pack["meta"]["pack_date"] == "2026-05-01"
            assert pack["meta"]["pack_date_method"] == "cli_override"

    def test_pack_date_heuristic_uses_last_month(self):
        """Without CLI arg, pack_date should be last-covered-month-end + 15 days."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "test_pack.xlsx"
            _make_pack_workbook(
                xlsx, num_months=3, last_month=datetime(2026, 3, 31)
            )

            pack = _parser_mod.parse_investor_pack(xlsx)

            # 2026-02-28 (day=28 written by helper) + 15 days = 2026-03-15
            # Actually the helper uses day=28 for safety; last month written
            # will be 2026-03-28; + 15 = 2026-04-12
            assert pack["meta"]["pack_date"].startswith("2026-04-")
            assert "heuristic" in pack["meta"]["pack_date_method"]
            assert pack["meta"]["data_range"]["month_count"] == 3
            assert pack["meta"]["data_range"]["last_month"] == "2026-03"

    def test_invalid_cli_pack_date_raises(self):
        """Bad --pack-date format → ValueError at parse time."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "test_pack.xlsx"
            _make_pack_workbook(xlsx, num_months=3)

            with pytest.raises(ValueError, match="must be YYYY-MM-DD"):
                _parser_mod.parse_investor_pack(xlsx, cli_pack_date="April 15 2026")

    def test_deterministic_output(self):
        """Same input → same JSON bytes. No timestamps in content (except parsed_at_utc
        which is intentionally a runtime metadata field, ignored for determinism check)."""
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = Path(tmp) / "test_pack.xlsx"
            _make_pack_workbook(xlsx, num_months=3)

            p1 = _parser_mod.parse_investor_pack(xlsx, cli_pack_date="2026-04-15")
            p2 = _parser_mod.parse_investor_pack(xlsx, cli_pack_date="2026-04-15")

            # Strip parsed_at_utc (not content-deterministic — that's runtime metadata)
            for p in (p1, p2):
                p["meta"].pop("parsed_at_utc", None)

            j1 = json.dumps(p1, sort_keys=True, default=str)
            j2 = json.dumps(p2, sort_keys=True, default=str)
            assert j1 == j2


# ── Enrichment tests ──────────────────────────────────────────────────────────


class TestQuarterlyPackEnrichment:
    """core/analysis_tamara.py merges the latest pack under ``quarterly_pack``."""

    def test_missing_pack_folder_is_noop(self, tmp_path):
        """A Tamara snapshot with no investor_packs sibling → no quarterly_pack key.
        Existing deployments without the new pipeline must keep working."""
        from core.analysis_tamara import _find_latest_investor_pack

        # Simulate data/Tamara/KSA/2026-04-09_tamara_ksa.json with no sibling folder
        (tmp_path / "Tamara" / "KSA").mkdir(parents=True)
        fake_snapshot = tmp_path / "Tamara" / "KSA" / "2026-04-09_tamara_ksa.json"
        fake_snapshot.write_text("{}")

        assert _find_latest_investor_pack(str(fake_snapshot)) is None

    def test_finds_latest_pack_by_filename_sort(self, tmp_path):
        """Multiple packs in the folder → latest filename-date wins."""
        from core.analysis_tamara import _find_latest_investor_pack

        (tmp_path / "Tamara" / "KSA").mkdir(parents=True)
        packs_dir = tmp_path / "Tamara" / "investor_packs"
        packs_dir.mkdir()
        (packs_dir / "2026-01-15_investor_pack.json").write_text("{}")
        (packs_dir / "2026-04-15_investor_pack.json").write_text("{}")
        (packs_dir / "2026-02-15_investor_pack.json").write_text("{}")

        fake_snapshot = tmp_path / "Tamara" / "KSA" / "2026-04-09_tamara_ksa.json"
        fake_snapshot.write_text("{}")

        result = _find_latest_investor_pack(str(fake_snapshot))
        assert result is not None
        assert result.endswith("2026-04-15_investor_pack.json")

    def test_mom_delta_computation(self):
        """_mom_delta returns the right structure and signs."""
        from core.analysis_tamara import _mom_delta

        series = {"2026-01": 100, "2026-02": 110, "2026-03": 99}
        delta = _mom_delta(series, ["2026-01", "2026-02", "2026-03"])

        assert delta["latest"] == 99
        assert delta["latest_month"] == "2026-03"
        assert delta["prior"] == 110
        assert delta["prior_month"] == "2026-02"
        assert delta["abs_delta"] == -11
        assert delta["pct_delta"] == pytest.approx(-0.1)  # -10%

    def test_mom_delta_handles_missing_values(self):
        """_mom_delta returns None deltas when prior or latest is missing."""
        from core.analysis_tamara import _mom_delta

        assert _mom_delta({}, [])["latest"] is None
        assert _mom_delta({"2026-03": 99}, ["2026-03"])["latest"] == 99
        assert _mom_delta({"2026-03": 99}, ["2026-03"])["prior"] is None

        # Prior is None → pct_delta stays None
        delta = _mom_delta(
            {"2026-02": None, "2026-03": 99}, ["2026-02", "2026-03"]
        )
        assert delta["pct_delta"] is None

    def test_enrichment_adds_quarterly_pack_key(self, tmp_path):
        """End-to-end: build a pack file via the parser, then parse_tamara_data
        should find it and attach a quarterly_pack key."""
        import json as _json

        from core.analysis_tamara import parse_tamara_data

        # Build the directory layout parse_tamara_data expects
        ksa_dir = tmp_path / "Tamara" / "KSA"
        ksa_dir.mkdir(parents=True)
        packs_dir = tmp_path / "Tamara" / "investor_packs"
        packs_dir.mkdir()

        # Minimal valid snapshot (parse_tamara_data calls enrichers that read
        # various keys — empty dict is OK, they all use .get defaults)
        snapshot_path = ksa_dir / "2026-04-09_tamara_ksa.json"
        snapshot_path.write_text(_json.dumps({}))

        # Build a valid pack + write its JSON to the packs folder
        pack_xlsx = tmp_path / "pack.xlsx"
        _make_pack_workbook(pack_xlsx, num_months=3)
        pack = _parser_mod.parse_investor_pack(pack_xlsx, cli_pack_date="2026-04-15")
        (packs_dir / "2026-04-15_investor_pack.json").write_text(
            _json.dumps(pack, default=str)
        )

        # Run parse_tamara_data → should merge the pack in
        result = parse_tamara_data(str(snapshot_path))

        assert "quarterly_pack" in result
        qp = result["quarterly_pack"]
        assert qp["pack_date"] == "2026-04-15"
        # Headline FS (cons) should have a GMV entry
        assert "Total GMV" in qp["headline_fs"]["cons"]
        # MoM delta shape
        delta = qp["headline_fs"]["cons"]["Total GMV"]
        assert delta["latest_month"] == "2026-03"
        assert delta["latest"] is not None
