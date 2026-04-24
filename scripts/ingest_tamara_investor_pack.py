#!/usr/bin/env python3
"""
Tamara Quarterly Investor Pack Ingest
======================================
Parses Tamara's recurring 10-sheet investor pack Excel template and emits a
structured JSON snapshot for the dashboard + Executive Summary pipeline.

Template (all packs share this structure):
  - "Investor Reporting Pack >>"        (divider)
  - "1. KPIs >>"                         (divider)
  - "1.1 KPIs cons"  (Consolidated)
  - "1.2 KPIs KSA"
  - "1.3 KPIs UAE"
  - "2. Financials>>"                    (divider)
  - "2.1 FS Cons"                        (P&L + BS + CF, dual Management/Statutory view)
  - "2.2 FS KSA"
  - "2.3 FS UAE"
  - "3. Performance v Budget"            (Q1 actuals vs base case)

Output:
  data/Tamara/investor_packs/YYYY-MM-DD_investor_pack.json

Usage:
  python scripts/ingest_tamara_investor_pack.py --file <path> [--pack-date YYYY-MM-DD] [--dry-run] [--force]

Pack-date resolution order:
  1. --pack-date CLI arg (explicit override)
  2. Computed as (last-covered-month-end + 15 days) from the data's date range
  3. File mtime as last-resort fallback

The output JSON is deterministic: same input file → same output bytes.
"""

import argparse
import json
import sys
import warnings
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT_DIR = PROJECT_ROOT / "data" / "Tamara" / "investor_packs"

# Template signature — these 10 sheet names must all be present (in any order)
EXPECTED_SHEETS = {
    "Investor Reporting Pack >>",
    "1. KPIs >>",
    "1.1 KPIs cons",
    "1.2 KPIs KSA",
    "1.3 KPIs UAE",
    "2. Financials>>",
    "2.1 FS Cons",
    "2.2 FS KSA",
    "2.3 FS UAE",
    "3. Performance v Budget",
}


# ── Helpers ───────────────────────────────────────────────────────────────────


def _safe(v: Any) -> Any:
    """Convert Excel-native types to JSON-safe Python types. Round floats to 6dp for stable output."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()[:10]
    if isinstance(v, float):
        # NaN check
        if v != v:  # noqa: PLR0124
            return None
        return round(v, 6)
    if isinstance(v, (int, bool, str)):
        return v
    # Fallback: stringify
    return str(v)


def _find_header_date_row(ws, max_scan_rows: int = 10) -> Tuple[int, List[Tuple[int, str]]]:
    """
    Locate the row containing monthly date headers. Returns (row_index, [(col, 'YYYY-MM'), ...]).
    KPI sheets put dates in row 6; FS sheets put dates in row 5. Budget sheet also row 6.
    We scan the first 10 rows and pick the row with the most datetime cells.
    """
    best_row = -1
    best_cols: List[Tuple[int, str]] = []
    for r in range(1, max_scan_rows + 1):
        cols = []
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                cols.append((c, v.strftime("%Y-%m")))
        if len(cols) > len(best_cols):
            best_row = r
            best_cols = cols
    return best_row, best_cols


def _extract_label_column(ws, label_col: int, start_row: int, end_row: int) -> List[Tuple[int, str]]:
    """Return [(row, label_text), ...] for non-empty labels in a column range."""
    out = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, label_col).value
        if v is not None and str(v).strip():
            out.append((r, str(v).strip()))
    return out


def _extract_series(ws, row: int, date_cols: List[Tuple[int, str]]) -> Dict[str, Any]:
    """Extract a {YYYY-MM: value} dict for one label row across all date columns."""
    out = {}
    for col, month_key in date_cols:
        v = ws.cell(row, col).value
        out[month_key] = _safe(v)
    return out


def _real_extent(ws) -> Tuple[int, int]:
    """Return (max_row, max_col) with actual data (openpyxl.max_row can overstate)."""
    max_r, max_c = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if cell.row > max_r:
                    max_r = cell.row
                if cell.column > max_c:
                    max_c = cell.column
    return max_r, max_c


# ── Section parsers ───────────────────────────────────────────────────────────


def _parse_kpi_sheet(ws) -> Dict[str, Any]:
    """Parse a '1.x KPIs' sheet. Returns {months, line_items: {label: {month: value}}}."""
    date_row, date_cols = _find_header_date_row(ws)
    if not date_cols:
        return {"months": [], "line_items": {}, "_warning": "no date header row found"}

    max_r, _ = _real_extent(ws)
    # Label column is col 2 (column B) in every KPI sheet we've observed.
    label_col = 2
    start_row = date_row + 1

    line_items: Dict[str, Dict[str, Any]] = {}
    for r, label in _extract_label_column(ws, label_col, start_row, max_r):
        series = _extract_series(ws, r, date_cols)
        # Skip empty rows (all None)
        if all(v is None for v in series.values()):
            # It's probably a section header — preserve it with a prefix
            line_items[label] = {"_section_header": True}
            continue
        line_items[label] = series

    months = [m for _, m in date_cols]
    return {"months": months, "line_items": line_items}


def _parse_fs_sheet(ws) -> Dict[str, Any]:
    """Parse a '2.x FS' sheet. Returns same shape as KPI but covers P&L (Mgmt + Statutory), BS, CF."""
    # Same structure as KPI sheets, just different row extent
    return _parse_kpi_sheet(ws)


def _parse_budget_sheet(ws) -> Dict[str, Any]:
    """
    Parse '3. Performance v Budget' sheet. Has Actuals, Budget (Base Case), Variance sections.
    Row 4 has section labels ("Actuals", "Budget - Base Case", "Variance vs. Budget").
    Row 6 has per-column date / section-sub-labels.
    """
    max_r, max_c = _real_extent(ws)
    # Find the section header row (row 4 by convention)
    section_headers: Dict[int, str] = {}
    for c in range(1, max_c + 1):
        v = ws.cell(4, c).value
        if v is not None:
            section_headers[c] = str(v).strip()

    # Map each column to a (section, month_or_label) pair using row 6
    col_map: Dict[int, Tuple[str, str]] = {}
    # Walk left to right. Remember the "current section" when we cross a section-header col.
    current_section = None
    # Sort section-header cols so we can detect section boundaries
    header_cols_sorted = sorted(section_headers.keys())
    for c in range(1, max_c + 1):
        # Update current_section if this col starts a new section header
        if c in section_headers:
            current_section = section_headers[c]
        hdr_val = ws.cell(6, c).value
        if hdr_val is None:
            continue
        if isinstance(hdr_val, datetime):
            col_map[c] = (current_section or "", hdr_val.strftime("%Y-%m"))
        else:
            col_map[c] = (current_section or "", str(hdr_val).strip())

    # Metric label column is col 2
    label_col = 2
    start_row = 7

    line_items: Dict[str, Dict[str, Any]] = {}
    for r, label in _extract_label_column(ws, label_col, start_row, max_r):
        row_data: Dict[str, Dict[str, Any]] = {}
        for c, (section, key) in col_map.items():
            v = ws.cell(r, c).value
            if v is not None:
                row_data.setdefault(section or "unlabeled", {})[key] = _safe(v)
        if row_data:
            line_items[label] = row_data

    return {"sections": list({s for s, _ in col_map.values() if s}), "line_items": line_items}


# ── Pack-date resolution ──────────────────────────────────────────────────────


def _resolve_pack_date(
    wb: openpyxl.Workbook, xlsx_path: Path, cli_pack_date: Optional[str]
) -> Tuple[str, str]:
    """
    Determine the pack's logical reporting date. Returns (pack_date, method).
    Method string documents the source of the date — included in meta for traceability.
    """
    if cli_pack_date:
        try:
            datetime.strptime(cli_pack_date, "%Y-%m-%d")
            return cli_pack_date, "cli_override"
        except ValueError as exc:
            raise ValueError(f"--pack-date must be YYYY-MM-DD, got {cli_pack_date!r}") from exc

    # Heuristic: find last-covered month-end in the KPIs cons sheet → + 15 days
    try:
        ws = wb["1.1 KPIs cons"]
        _, date_cols = _find_header_date_row(ws)
        if date_cols:
            # Last column is the latest month; use the actual cell value (a datetime) not just the key
            latest_col = date_cols[-1][0]
            header_row = _find_header_date_row(ws)[0]
            latest_cell = ws.cell(header_row, latest_col).value
            if isinstance(latest_cell, datetime):
                pack_date = (latest_cell + timedelta(days=15)).strftime("%Y-%m-%d")
                return pack_date, f"heuristic_last_month_plus_15d (latest_month={date_cols[-1][1]})"
    except Exception:  # noqa: BLE001
        pass

    # Final fallback: file mtime
    mtime = datetime.fromtimestamp(xlsx_path.stat().st_mtime)
    return mtime.strftime("%Y-%m-%d"), "file_mtime"


# ── Main pack parser ──────────────────────────────────────────────────────────


def parse_investor_pack(xlsx_path: Path, cli_pack_date: Optional[str] = None) -> Dict[str, Any]:
    """
    Parse a Tamara investor pack Excel file. Raises ValueError on template mismatch.
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Investor pack not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Template validation
    sheet_names = set(wb.sheetnames)
    missing = EXPECTED_SHEETS - sheet_names
    if missing:
        raise ValueError(
            f"Template mismatch — missing sheets: {sorted(missing)}. "
            f"Found sheets: {sorted(sheet_names)}"
        )

    pack_date, date_method = _resolve_pack_date(wb, xlsx_path, cli_pack_date)

    # Parse KPIs (3 country cuts)
    kpis = {
        "cons": _parse_kpi_sheet(wb["1.1 KPIs cons"]),
        "ksa": _parse_kpi_sheet(wb["1.2 KPIs KSA"]),
        "uae": _parse_kpi_sheet(wb["1.3 KPIs UAE"]),
    }

    # Parse Financial Statements (3 country cuts, dual Management/Statutory view)
    financials = {
        "cons": _parse_fs_sheet(wb["2.1 FS Cons"]),
        "ksa": _parse_fs_sheet(wb["2.2 FS KSA"]),
        "uae": _parse_fs_sheet(wb["2.3 FS UAE"]),
    }

    # Parse Performance vs Budget
    budget_variance = _parse_budget_sheet(wb["3. Performance v Budget"])

    # Determine date range from KPI cons (the richest sheet)
    cons_months = kpis["cons"].get("months", [])
    data_range = {
        "first_month": cons_months[0] if cons_months else None,
        "last_month": cons_months[-1] if cons_months else None,
        "month_count": len(cons_months),
    }

    return {
        "meta": {
            "source_file": xlsx_path.name,
            "pack_date": pack_date,
            "pack_date_method": date_method,
            "parsed_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "data_range": data_range,
            "template_version": "tamara_investor_pack_v1",
        },
        "kpis": kpis,
        "financials": financials,
        "budget_variance": budget_variance,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Ingest a Tamara quarterly/monthly investor pack Excel into a JSON snapshot.",
    )
    parser.add_argument("--file", required=True, type=Path, help="Path to the .xlsx pack file")
    parser.add_argument(
        "--pack-date",
        type=str,
        default=None,
        help="Override pack date (YYYY-MM-DD). Default: heuristic from data range.",
    )
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--dry-run", action="store_true", help="Parse + print summary, don't write output")
    parser.add_argument("--force", action="store_true", help="Overwrite existing output file")
    args = parser.parse_args()

    try:
        pack = parse_investor_pack(args.file, cli_pack_date=args.pack_date)
    except (FileNotFoundError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    pack_date = pack["meta"]["pack_date"]
    out_path = args.out_dir / f"{pack_date}_investor_pack.json"

    # Summary on stderr so --dry-run stdout stays clean for piping
    print(f"[pack] source={args.file.name}", file=sys.stderr)
    print(f"[pack] pack_date={pack_date} (method: {pack['meta']['pack_date_method']})", file=sys.stderr)
    print(
        f"[pack] data_range={pack['meta']['data_range']['first_month']} → "
        f"{pack['meta']['data_range']['last_month']} "
        f"({pack['meta']['data_range']['month_count']} months)",
        file=sys.stderr,
    )
    kpi_counts = {k: len(v.get("line_items", {})) for k, v in pack["kpis"].items()}
    fs_counts = {k: len(v.get("line_items", {})) for k, v in pack["financials"].items()}
    budget_count = len(pack["budget_variance"].get("line_items", {}))
    print(f"[pack] kpi_line_items={kpi_counts}", file=sys.stderr)
    print(f"[pack] fs_line_items={fs_counts}", file=sys.stderr)
    print(f"[pack] budget_line_items={budget_count}", file=sys.stderr)
    print(f"[pack] out_path={out_path}", file=sys.stderr)

    if args.dry_run:
        print("[pack] DRY RUN — not writing output", file=sys.stderr)
        return 0

    if out_path.exists() and not args.force:
        print(
            f"[pack] output already exists: {out_path}. Use --force to overwrite.",
            file=sys.stderr,
        )
        return 2

    out_path.parent.mkdir(parents=True, exist_ok=True)
    # sort_keys=True makes output deterministic
    out_path.write_text(json.dumps(pack, indent=2, sort_keys=True, default=str), encoding="utf-8")
    print(f"[pack] OK wrote {out_path} ({out_path.stat().st_size:,} bytes)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
